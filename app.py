"""
MSA Scholarship Matching System - Admin Application
"""

import os
import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for, jsonify, session, flash, send_file
)
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from config import config
from models import db, Scholarship, Student, Application, Score, Award, ApplicationFile
from scoring import score_student
from matcher import get_eligible_scholarships, suggest_awards
from notifications import send_award_notification

app = Flask(__name__)
env = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[env])

# Initialize database
db.init_app(app)

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


SAMPLE_SCHOLARSHIPS = [
    {
        'sort_name': 'ACADEMIC EXCELLENCE',
        'official_name': 'Academic Excellence Scholarship',
        'account_number': 'ACC-001',
        'ug_eligible': True, 'ms_eligible': True,
        'fall_award': True, 'spring_award': True,
        'available_amount': Decimal('30000'),
        'restrictions': 'Minimum 3.5 GPA required',
        'notes': 'Awarded to top academic performers'
    },
    {
        'sort_name': 'MERIT AWARD',
        'official_name': 'Merit Scholarship Award',
        'account_number': 'ACC-002',
        'ug_eligible': True, 'ms_eligible': False,
        'fall_award': True, 'spring_award': False,
        'available_amount': Decimal('20000'),
        'restrictions': 'Undergraduate students only',
        'notes': 'Recognizes outstanding undergraduate achievement'
    },
    {
        'sort_name': 'GRADUATE LEADERSHIP',
        'official_name': 'Graduate Leadership Scholarship',
        'account_number': 'ACC-003',
        'ug_eligible': False, 'ms_eligible': True,
        'fall_award': True, 'spring_award': True,
        'available_amount': Decimal('25000'),
        'restrictions': 'Graduate students only; demonstrated leadership required',
        'notes': 'For students with strong leadership backgrounds'
    },
    {
        'sort_name': 'FINANCIAL NEED',
        'official_name': 'Financial Need Assistance Grant',
        'account_number': 'ACC-004',
        'ug_eligible': True, 'ms_eligible': True,
        'fall_award': True, 'spring_award': True,
        'available_amount': Decimal('40000'),
        'restrictions': 'High financial need required',
        'notes': 'Priority given to students with high financial need'
    },
    {
        'sort_name': 'PROFESSIONAL DEVELOPMENT',
        'official_name': 'Professional Development Scholarship',
        'account_number': 'ACC-005',
        'ug_eligible': False, 'ms_eligible': True,
        'fall_award': False, 'spring_award': True,
        'available_amount': Decimal('15000'),
        'restrictions': 'Must have internship or work experience',
        'notes': 'For students with strong professional backgrounds'
    },
    {
        'sort_name': 'COMMUNITY SERVICE',
        'official_name': 'Community Service Award',
        'account_number': 'ACC-006',
        'ug_eligible': True, 'ms_eligible': True,
        'fall_award': True, 'spring_award': True,
        'available_amount': Decimal('10000'),
        'restrictions': 'Demonstrated community involvement required',
        'notes': 'Recognizes students with strong extracurricular contributions'
    },
]


def seed_sample_data():
    """Seed database with sample scholarships when no real data files are present."""
    if Scholarship.query.count() > 0:
        return  # Already has data, skip seeding

    print("No scholarships found. Seeding sample data for demo...")
    for data in SAMPLE_SCHOLARSHIPS:
        scholarship = Scholarship(**data)
        db.session.add(scholarship)
    db.session.commit()
    print(f"Seeded {len(SAMPLE_SCHOLARSHIPS)} sample scholarships.")


def init_db():
    """Initialize database with tables."""
    with app.app_context():
        db.create_all()
        _migrate_notified_to_approved()
        print("Database initialized!")


def _migrate_notified_to_approved():
    """Convert legacy 'notified' award status back to 'approved'.
    The email tracking is handled by email_status field instead."""
    try:
        notified = Award.query.filter_by(status='notified').all()
        for award in notified:
            award.status = 'approved'
        if notified:
            db.session.commit()
            print(f"Migrated {len(notified)} 'notified' awards → 'approved'")
    except Exception:
        db.session.rollback()


def load_scholarships_from_excel():
    """Load scholarships from master Excel file."""
    from openpyxl import load_workbook

    excel_file = app.config['SCHOLARSHIP_MASTER_FILE']
    available_funds_file = app.config['AVAILABLE_FUNDS_FILE']

    if not os.path.exists(excel_file) or not os.path.exists(available_funds_file):
        print(f"Warning: Excel files not found")
        return 0

    # Load fund information
    wb = load_workbook(available_funds_file)
    ws = wb.active

    funds_dict = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:  # Sort name
            sort_name = row[0].strip()
            available = row[7]  # COMMITTEE WORKING TOTAL column
            if available:
                funds_dict[sort_name] = Decimal(str(available))

    # Load scholarship details
    wb = load_workbook(excel_file)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # Official name
            sort_name = row[1] if row[0] is None else (row[0] or row[1])

            existing = Scholarship.query.filter_by(sort_name=sort_name).first()
            if existing:
                continue

            scholarship = Scholarship(
                sort_name=str(sort_name).strip(),
                official_name=str(row[1]).strip() if row[1] else str(sort_name).strip(),
                account_number=str(row[2]) if row[2] else None,
                ug_eligible=_parse_level(row[2], 'ug'),
                ms_eligible=_parse_level(row[2], 'ms'),
                restrictions=str(row[4]) if row[4] else '',
                notes=str(row[9]) if row[9] else '',
                available_amount=funds_dict.get(str(sort_name).strip(), Decimal('0'))
            )

            db.session.add(scholarship)
            count += 1

    db.session.commit()
    print(f"Loaded {count} scholarships from Excel")
    return count


def _parse_level(text, level):
    """Check if text indicates UG or MS eligibility."""
    if not text:
        return True  # Default to both if not specified
    text_lower = str(text).lower()
    if level == 'ug':
        return 'undergrad' in text_lower or 'ug' in text_lower or 'undergraduate' in text_lower
    else:
        return 'grad' in text_lower or 'ms' in text_lower or 'master' in text_lower


def login_required(f):
    """Simple login requirement."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Admin-only access."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or session.get('role') != 'admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def evaluator_required(f):
    """Evaluator-only access."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or session.get('role') != 'evaluator':
            flash('Access denied. Evaluator privileges required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


# Routes

@app.route('/')
def index():
    """Home/dashboard route."""
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    role = session.get('role')
    
    # Show appropriate dashboard based on role
    if role == 'admin':
        # Admin dashboard - management focused
        stats = {
            'total_scholarships': Scholarship.query.count(),
            'total_students': Student.query.count(),
            'total_awards': Award.query.count(),
            'pending_approvals': Award.query.filter_by(status='pending').count(),
            'approved_awards': Award.query.filter(Award.status.in_(['approved', 'notified'])).count(),
        }
        return render_template('admin_dashboard.html', stats=stats, role=role)
    else:
        # Evaluator dashboard - program execution focused
        stats = {
            'total_students': Student.query.count(),
            'total_awards': Award.query.count(),
            'pending_approvals': Award.query.filter_by(status='pending').count(),
            'approved_awards': Award.query.filter(Award.status.in_(['approved', 'notified'])).count(),
        }
        return render_template('dashboard.html', stats=stats, role=role)


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login route with role selection."""
    if request.method == 'POST':
        role = request.form.get('role')  # 'admin' or 'evaluator'
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        
        # Define test credentials for different roles
        evaluator_credentials = {
            'username': os.environ.get('EVALUATOR_USERNAME', 'evaluator@msu.edu'),
            'password': os.environ.get('EVALUATOR_PASSWORD', 'evaluator123')
        }
        
        admin_credentials = {
            'username': os.environ.get('ADMIN_USERNAME', 'admin@msu.edu'),
            'password': os.environ.get('ADMIN_PASSWORD', 'admin123')
        }
        
        # Validate credentials based on role
        if role == 'evaluator':
            if username == evaluator_credentials['username'] and password == evaluator_credentials['password']:
                session['logged_in'] = True
                session['role'] = 'evaluator'
                session['username'] = username
                session['login_time'] = datetime.utcnow().isoformat()
                return redirect(url_for('index'))
            else:
                flash('Invalid evaluator credentials', 'error')
        
        elif role == 'admin':
            if username == admin_credentials['username'] and password == admin_credentials['password']:
                session['logged_in'] = True
                session['role'] = 'admin'
                session['username'] = username
                session['login_time'] = datetime.utcnow().isoformat()
                return redirect(url_for('index'))
            else:
                flash('Invalid administrator credentials', 'error')
        
        else:
            flash('Invalid role selected', 'error')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logout route."""
    session.clear()
    return redirect(url_for('login'))


# ==================== SCHOLARSHIP MANAGEMENT ROUTES ====================

@app.route('/scholarships', methods=['GET'])
@admin_required
def manage_scholarships():
    """View all scholarships."""
    scholarships = Scholarship.query.all()
    
    # Calculate allocation info for each scholarship
    for scholarship in scholarships:
        scholarship.allocated = sum(a.amount for a in scholarship.awards if a.status in ('approved', 'notified'))
        scholarship.remaining = scholarship.available_amount - scholarship.allocated
    
    return render_template('scholarships.html', scholarships=scholarships)


@app.route('/scholarships/add', methods=['GET', 'POST'])
@admin_required
def add_scholarship():
    """Add a new scholarship."""
    if request.method == 'POST':
        try:
            sort_name = request.form.get('sort_name', '').strip()
            official_name = request.form.get('official_name', '').strip()
            account_number = request.form.get('account_number', '').strip()
            available_amount = Decimal(request.form.get('available_amount', '0'))
            ug_eligible = request.form.get('ug_eligible') == 'on'
            ms_eligible = request.form.get('ms_eligible') == 'on'
            fall_award = request.form.get('fall_award') == 'on'
            spring_award = request.form.get('spring_award') == 'on'
            restrictions = request.form.get('restrictions', '').strip()
            notes = request.form.get('notes', '').strip()

            # Validate
            if not sort_name or not official_name:
                flash('Sort name and official name are required', 'error')
                return redirect(url_for('add_scholarship'))

            if Scholarship.query.filter_by(sort_name=sort_name).first():
                flash('A scholarship with this sort name already exists', 'error')
                return redirect(url_for('add_scholarship'))

            scholarship = Scholarship(
                sort_name=sort_name,
                official_name=official_name,
                account_number=account_number if account_number else None,
                available_amount=available_amount,
                ug_eligible=ug_eligible,
                ms_eligible=ms_eligible,
                fall_award=fall_award,
                spring_award=spring_award,
                restrictions=restrictions,
                notes=notes
            )
            db.session.add(scholarship)
            db.session.commit()
            flash(f'Scholarship "{sort_name}" added successfully!', 'success')
            return redirect(url_for('manage_scholarships'))
        except Exception as e:
            flash(f'Error adding scholarship: {str(e)}', 'error')
            return redirect(url_for('add_scholarship'))

    return render_template('add_scholarship.html')


@app.route('/scholarships/edit/<int:scholarship_id>', methods=['GET', 'POST'])
@admin_required
def edit_scholarship(scholarship_id):
    """Edit an existing scholarship."""
    scholarship = Scholarship.query.get_or_404(scholarship_id)

    if request.method == 'POST':
        try:
            scholarship.sort_name = request.form.get('sort_name', '').strip()
            scholarship.official_name = request.form.get('official_name', '').strip()
            scholarship.account_number = request.form.get('account_number', '').strip() or None
            scholarship.available_amount = Decimal(request.form.get('available_amount', '0'))
            scholarship.ug_eligible = request.form.get('ug_eligible') == 'on'
            scholarship.ms_eligible = request.form.get('ms_eligible') == 'on'
            scholarship.fall_award = request.form.get('fall_award') == 'on'
            scholarship.spring_award = request.form.get('spring_award') == 'on'
            scholarship.restrictions = request.form.get('restrictions', '').strip()
            scholarship.notes = request.form.get('notes', '').strip()
            scholarship.updated_at = datetime.utcnow()

            db.session.commit()
            flash(f'Scholarship "{scholarship.sort_name}" updated successfully!', 'success')
            return redirect(url_for('manage_scholarships'))
        except Exception as e:
            flash(f'Error updating scholarship: {str(e)}', 'error')

    return render_template('edit_scholarship.html', scholarship=scholarship)


@app.route('/scholarships/delete/<int:scholarship_id>', methods=['POST'])
@admin_required
def delete_scholarship(scholarship_id):
    """Delete a scholarship."""
    scholarship = Scholarship.query.get_or_404(scholarship_id)
    
    # Check if scholarship has awards
    if scholarship.awards:
        flash('Cannot delete scholarship with existing awards', 'error')
        return redirect(url_for('manage_scholarships'))
    
    try:
        sort_name = scholarship.sort_name
        db.session.delete(scholarship)
        db.session.commit()
        flash(f'Scholarship "{sort_name}" deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting scholarship: {str(e)}', 'error')
    
    return redirect(url_for('manage_scholarships'))


# ==================== FILE UPLOAD ROUTES ====================

@app.route('/upload', methods=['GET', 'POST'])
@evaluator_required
def upload_applications():
    """Upload applications CSV or Excel."""

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file provided', 'error')
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        # Check file format
        file_ext = None
        if file.filename.endswith('.csv'):
            file_ext = 'csv'
        elif file.filename.endswith(('.xlsx', '.xls')):
            file_ext = 'excel'
        else:
            flash('Please upload a CSV or Excel file (.csv, .xlsx, .xls)', 'error')
            return redirect(request.url)

        # Save file
        filename = f"applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        app_file = ApplicationFile(
            filename=filename,
            upload_path=filepath,
            uploaded_by=session.get('admin_user', 'system')
        )
        db.session.add(app_file)
        db.session.flush()

        file.save(filepath)

        # Process file
        try:
            if file_ext == 'csv':
                num_processed = _process_applications_csv(filepath)
            else:
                num_processed = _process_applications_excel(filepath)
            
            app_file.status = 'success'
            app_file.num_students = num_processed
            app_file.processed_at = datetime.utcnow()
            db.session.commit()
            flash(f'Successfully uploaded and processed {num_processed} applications', 'success')
            return redirect(url_for('review_matches'))
        except Exception as e:
            app_file.status = 'error'
            app_file.error_message = str(e)
            db.session.commit()
            flash(f'Error processing file: {str(e)}', 'error')

    return render_template('upload.html')


@app.route('/review', methods=['GET'])
@evaluator_required
def review_matches():
    """Review and approve suggested scholarship matches."""

    students = Student.query.all()
    matches = []

    for student in students:
        score = Score.query.filter_by(student_id=student.id).first()
        if not score:
            continue

        eligible = get_eligible_scholarships(student)
        awards = Award.query.filter_by(student_id=student.id).all()

        matches.append({
            'student': student,
            'score': score,
            'eligible_scholarships': eligible,
            'current_awards': awards
        })

    # Sort by score descending
    matches.sort(key=lambda x: x['score'].total_score, reverse=True)

    return render_template('review.html', matches=matches)


@app.route('/approve', methods=['GET', 'POST'])
@evaluator_required
def approve_awards():
    """Manager approves/rejects and sets amounts for awards."""

    if request.method == 'POST':
        data = request.get_json()
        award_id = data.get('award_id')
        action = data.get('action')  # 'approve' or 'reject'
        amount = data.get('amount')

        award = Award.query.get(award_id)
        if not award:
            return jsonify({'error': 'Award not found'}), 404

        if action == 'approve':
            try:
                # Clean the amount value - remove $ and commas
                if isinstance(amount, str):
                    amount = amount.replace('$', '').replace(',', '').strip()
                
                award.status = 'approved'
                award.amount = Decimal(str(amount))
                award.approved_by = session.get('admin_user', 'manager')
                award.approved_at = datetime.utcnow()
                award.email_status = 'not_sent'  # Reset email status when approving
                db.session.commit()
                return jsonify({'success': True, 'status': 'approved'})
            except Exception as e:
                return jsonify({'error': f'Error approving award: {str(e)}'}), 400

        elif action == 'reject':
            try:
                award.status = 'rejected'
                db.session.commit()
                return jsonify({'success': True, 'status': 'rejected'})
            except Exception as e:
                return jsonify({'error': f'Error rejecting award: {str(e)}'}), 400

    # Get pending and approved awards
    pending_awards = Award.query.filter_by(status='pending').all()
    approved_awards = Award.query.filter(Award.status.in_(['approved', 'notified'])).all()

    # Students with scores but NO award at all (not pending, approved, or rejected)
    all_awarded_ids = db.session.query(Award.student_id).distinct()
    unmatched_students = Student.query.filter(
        ~Student.id.in_(all_awarded_ids)
    ).join(Score, Score.student_id == Student.id).all()

    all_scholarships = Scholarship.query.filter(Scholarship.available_amount > 0).all()

    return render_template('approve.html',
                         pending_awards=pending_awards,
                         approved_awards=approved_awards,
                         unmatched_students=unmatched_students,
                         all_scholarships=all_scholarships)


@app.route('/assign-award', methods=['POST'])
@evaluator_required
def assign_award():
    """Manually assign a scholarship to an unmatched student."""
    student_id = request.form.get('student_id')
    scholarship_id = request.form.get('scholarship_id')

    if not student_id or not scholarship_id:
        flash('Student and scholarship are required', 'error')
        return redirect(url_for('approve_awards'))

    existing = Award.query.filter_by(
        student_id=int(student_id),
        scholarship_id=int(scholarship_id)
    ).first()

    if existing:
        existing.status = 'pending'
        existing.amount = Decimal('5000')
    else:
        db.session.add(Award(
            student_id=int(student_id),
            scholarship_id=int(scholarship_id),
            amount=Decimal('5000'),
            status='pending'
        ))

    db.session.commit()
    flash('Award assigned successfully', 'success')
    return redirect(url_for('approve_awards'))


@app.route('/send-awards', methods=['POST'])
@evaluator_required
def send_awards():
    """Send award notification emails and log results."""

    approved_awards = Award.query.filter(
        Award.status.in_(['approved', 'notified']),
        Award.email_status == 'not_sent'
    ).all()

    results = {'sent': 0, 'failed': 0, 'messages': []}

    for award in approved_awards:
        try:
            result = send_award_notification(award, environment=app.config.get('MAIL_MODE', 'mock'))
            
            # Log the email attempt
            award.email_sent_to = award.student.email
            award.email_sent_at = datetime.utcnow()
            
            if result['success']:
                award.email_status = 'sent'
                results['sent'] += 1
                results['messages'].append({
                    'status': 'success',
                    'student': award.student.full_name(),
                    'message': f"✓ {award.student.full_name()}: Email sent successfully"
                })
            else:
                award.email_status = 'failed'
                award.email_error_message = result.get('error', 'Unknown error')
                results['failed'] += 1
                results['messages'].append({
                    'status': 'failed',
                    'student': award.student.full_name(),
                    'message': f"✗ {award.student.full_name()}: {result.get('error', 'Unknown error')}"
                })
        except Exception as e:
            award.email_status = 'failed'
            award.email_error_message = str(e)
            results['failed'] += 1
            results['messages'].append({
                'status': 'failed',
                'student': award.student.full_name(),
                'message': f"✗ {award.student.full_name()}: {str(e)}"
            })
        
        db.session.commit()

    return jsonify(results)


@app.route('/reports')
@login_required
def reports():
    """Scholarship and allocation reports."""

    scholarships = Scholarship.query.all()
    approved_awards = Award.query.filter(Award.status.in_(['approved', 'notified'])).all()
    total_distributed = sum((a.amount or 0) for a in approved_awards)

    # Pre-compute per-scholarship totals server-side to avoid lazy-load issues
    for s in scholarships:
        s._distributed = sum((a.amount or 0) for a in s.awards if a.status in ('approved', 'notified'))
        s._remaining = (s.available_amount or 0) - s._distributed

    report_data = {
        'scholarships': scholarships,
        'approved_awards': approved_awards,
        'total_distributed': total_distributed,
        'total_available': sum((s.available_amount or 0) for s in scholarships),
        'approved_count': Award.query.filter(Award.status.in_(['approved', 'notified'])).count(),
        'pending_count': Award.query.filter_by(status='pending').count(),
        'rejected_count': Award.query.filter_by(status='rejected').count(),
        'role': session.get('role')
    }

    return render_template('reports.html', **report_data)


@app.route('/reports/export-allocations', methods=['GET'])
@login_required
def export_allocations():
    """Export allocation report to Excel with student details and scholarship summary."""
    try:
        wb = openpyxl.Workbook()

        # ── Shared styles ──────────────────────────────────────────────────────
        title_fill   = PatternFill(start_color="0A3D2A", end_color="0A3D2A", fill_type="solid")
        header_fill  = PatternFill(start_color="0D5A3D", end_color="0D5A3D", fill_type="solid")
        sub_fill     = PatternFill(start_color="1A7A55", end_color="1A7A55", fill_type="solid")
        alt_fill     = PatternFill(start_color="F0F8F4", end_color="F0F8F4", fill_type="solid")
        title_font   = Font(bold=True, color="FFFFFF", size=14)
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        sub_font     = Font(bold=True, color="FFFFFF", size=10)
        bold_font    = Font(bold=True)
        currency_fmt = '"$"#,##0.00'
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        def style_header(cell, fill=header_fill, font=header_font):
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        def style_data(cell, alt=False, number_format=None, bold=False):
            cell.fill = alt_fill if alt else PatternFill()
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if number_format:
                cell.number_format = number_format
            if bold:
                cell.font = bold_font

        awards = Award.query.filter(Award.status.in_(['approved', 'notified'])).all()
        scholarships = Scholarship.query.all()
        total_allocated = sum((a.amount or 0) for a in awards)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — Student Award Details
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Student Award Details"

        # Title
        ws1.merge_cells('A1:I1')
        t = ws1['A1']
        t.value = "MSA Scholarship Program — Award Details"
        t.font = title_font
        t.fill = title_fill
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[1].height = 30

        ws1.merge_cells('A2:I2')
        ws1['A2'].value = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}   |   Total Recipients: {len(awards)}   |   Total Distributed: ${float(total_allocated):,.2f}"
        ws1['A2'].font = Font(italic=True, color="555555")
        ws1['A2'].alignment = Alignment(horizontal='center')
        ws1.row_dimensions[2].height = 18

        # Column headers
        headers = ['#', 'Student Name', 'Email', 'Program', 'GPA', 'Financial Need',
                   'Scholarship', 'Award Amount', 'Approved Date']
        for col, h in enumerate(headers, 1):
            style_header(ws1.cell(row=4, column=col))
            ws1.cell(row=4, column=col).value = h
        ws1.row_dimensions[4].height = 25

        # Data rows
        for i, award in enumerate(awards, 1):
            r = i + 4
            alt = (i % 2 == 0)
            need_labels = {'H': 'High', 'M': 'Medium', 'L': 'Low'}
            row_data = [
                i,
                award.student.full_name(),
                award.student.email or '',
                award.student.program_level,
                float(award.student.gpa) if award.student.gpa else '',
                need_labels.get(award.student.financial_need, 'N/A'),
                award.scholarship.official_name,
                float(award.amount),
                award.approved_at.strftime('%Y-%m-%d') if award.approved_at else '',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws1.cell(row=r, column=col, value=val)
                style_data(cell, alt=alt,
                           number_format=currency_fmt if col == 8 else None)
            ws1.row_dimensions[r].height = 18

        # Total row
        total_row = len(awards) + 5
        ws1.cell(row=total_row, column=7, value="TOTAL AWARDED").font = bold_font
        ws1.cell(row=total_row, column=7).alignment = Alignment(horizontal='right')
        tc = ws1.cell(row=total_row, column=8, value=float(total_allocated))
        tc.number_format = currency_fmt
        tc.font = Font(bold=True, color="0D5A3D", size=12)
        tc.border = thin_border

        # Column widths
        for col, w in zip('ABCDEFGHI', [4, 24, 28, 10, 7, 14, 34, 14, 14]):
            ws1.column_dimensions[col].width = w

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — Scholarship Summary
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Scholarship Summary")

        ws2.merge_cells('A1:F1')
        t2 = ws2['A1']
        t2.value = "MSA Scholarship Program — Fund Utilization by Scholarship"
        t2.font = title_font
        t2.fill = title_fill
        t2.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 30

        s_headers = ['Scholarship (Sort Name)', 'Official Name', 'Eligibility',
                     'Available Funds', 'Total Distributed', 'Remaining Balance']
        for col, h in enumerate(s_headers, 1):
            style_header(ws2.cell(row=3, column=col))
            ws2.cell(row=3, column=col).value = h
        ws2.row_dimensions[3].height = 25

        for i, s in enumerate(scholarships, 1):
            r = i + 3
            alt = (i % 2 == 0)
            distributed = sum((a.amount or 0) for a in s.awards if a.status in ('approved', 'notified'))
            remaining = (s.available_amount or 0) - distributed
            levels = []
            if s.ug_eligible: levels.append('UG')
            if s.ms_eligible: levels.append('MS')

            row_data = [
                s.sort_name,
                s.official_name,
                ' / '.join(levels) or 'N/A',
                float(s.available_amount or 0),
                float(distributed),
                float(remaining),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                style_data(cell, alt=alt,
                           number_format=currency_fmt if col in (4, 5, 6) else None)

        # Totals
        tot_r = len(scholarships) + 4
        ws2.cell(row=tot_r, column=3, value="TOTALS").font = bold_font
        ws2.cell(row=tot_r, column=3).alignment = Alignment(horizontal='right')
        for col, val in [(4, sum(float(s.available_amount or 0) for s in scholarships)),
                         (5, float(total_allocated)),
                         (6, sum(float(s.available_amount or 0) for s in scholarships) - float(total_allocated))]:
            c = ws2.cell(row=tot_r, column=col, value=val)
            c.number_format = currency_fmt
            c.font = Font(bold=True, color="0D5A3D")
            c.border = thin_border

        for col, w in zip('ABCDEF', [20, 36, 12, 16, 16, 16]):
            ws2.column_dimensions[col].width = w

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — Overview Summary
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Overview")

        ws3.merge_cells('A1:C1')
        t3 = ws3['A1']
        t3.value = "MSA Scholarship Program — Overview Summary"
        t3.font = title_font
        t3.fill = title_fill
        t3.alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[1].height = 30

        summary_rows = [
            ("Report Generated",    datetime.now().strftime('%B %d, %Y %I:%M %p'), None),
            ("", "", None),
            ("FUND SUMMARY", "", None),
            ("Total Available Funds", float(sum((s.available_amount or 0) for s in scholarships)), currency_fmt),
            ("Total Distributed",     float(total_allocated), currency_fmt),
            ("Remaining Balance",     float(sum((s.available_amount or 0) for s in scholarships)) - float(total_allocated), currency_fmt),
            ("Utilization Rate",      f"{(float(total_allocated)/float(sum((s.available_amount or 0) for s in scholarships))*100):.1f}%" if sum((s.available_amount or 0) for s in scholarships) > 0 else "N/A", None),
            ("", "", None),
            ("AWARD SUMMARY", "", None),
            ("Total Recipients",      len(awards), None),
            ("Scholarships Used",     len(set(a.scholarship_id for a in awards)), None),
            ("Pending Awards",        Award.query.filter_by(status='pending').count(), None),
            ("Rejected Awards",       Award.query.filter_by(status='rejected').count(), None),
        ]

        for i, (label, value, fmt) in enumerate(summary_rows, 3):
            lc = ws3.cell(row=i, column=1, value=label)
            vc = ws3.cell(row=i, column=2, value=value)
            if label in ("FUND SUMMARY", "AWARD SUMMARY"):
                style_header(lc, fill=sub_fill, font=sub_font)
                ws3.merge_cells(f'A{i}:C{i}')
            elif label:
                lc.font = bold_font
                lc.border = thin_border
                vc.border = thin_border
                if fmt:
                    vc.number_format = fmt

        ws3.column_dimensions['A'].width = 28
        ws3.column_dimensions['B'].width = 22

        # ── Save & send ────────────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'MSA_Scholarship_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'Error exporting allocations: {str(e)}', 'error')
        return redirect(url_for('reports'))


@app.route('/sample-format', methods=['GET'])
def download_sample_format():
    """Download sample application format as Excel file."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Applications"

        # Define styles
        header_fill = PatternFill(start_color="0D5A3D", end_color="0D5A3D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        instruction_fill = PatternFill(start_color="E8F4F0", end_color="E8F4F0", fill_type="solid")
        instruction_font = Font(italic=True, color="333333", size=9)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:M1')
        title = ws['A1']
        title.value = "MSA Scholarship Application Upload Format"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill(start_color="0A3D2A", end_color="0A3D2A", fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Instructions
        row = 3
        instructions = [
            "INSTRUCTIONS: Complete all fields below for each student. This template shows the required format.",
            "1. Student ID: Unique identifier for the student (can be ID number or first+last name)",
            "2. Names: Enter first, middle, and last names separately",
            "3. Email: Student's email address for contact",
            "4. GPA: Grade Point Average (e.g., 3.85)",
            "5. Program Level: Enter either 'UG' (Undergraduate) or 'MS' (Master's)",
            "6. Financial Need: Enter 'H' (High), 'M' (Medium), or 'L' (Low)",
            "7. Employer/Internship: Enter company names where applicable",
            "8. Activities/Experience: Describe relevant experience (can use commas to separate multiple entries)"
        ]

        for instruction in instructions:
            ws.merge_cells(f'A{row}:M{row}')
            cell = ws[f'A{row}']
            cell.value = instruction
            cell.font = instruction_font
            cell.fill = instruction_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1

        # Column headers
        headers = [
            'Student ID',
            'Student First Name',
            'Student Middle Name',
            'Student Last Name',
            'Email',
            'GPA',
            'Program Level',
            'Financial Need',
            'Permanent Position At',
            'Internship 2025 At',
            'Internship 2026 At',
            'Extracurricular Activities',
            'College Activities',
            'Work Experience'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        ws.row_dimensions[row].height = 40

        # Sample data row
        sample_data = [
            '20240001',
            'John',
            'Michael',
            'Smith',
            'john.smith@msu.edu',
            '3.85',
            'MS',
            'H',
            'Deloitte',
            'EY',
            'KPMG',
            'Student Government, Accounting Club, Volunteer Tutor',
            'Finance Committee Chair, Scholarship Recipient',
            'Auditor at Big 4 Firm (2 years), Bookkeeper at SMB (1 year)'
        ]

        row += 1
        for col, data in enumerate(sample_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = data
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.row_dimensions[row].height = 30

        # Add another blank row as template
        row += 2
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.row_dimensions[row].height = 30

        # Set column widths for better readability
        column_widths = {
            'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 20,
            'F': 8, 'G': 14, 'H': 12, 'I': 20, 'J': 16,
            'K': 16, 'L': 35, 'M': 30, 'N': 30
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='scholarship_application_sample.xlsx'
        )
    except Exception as e:
        flash(f'Error creating sample format: {str(e)}', 'error')
        return redirect(url_for('upload_applications'))


# Helper functions

def _process_applications_csv(filepath):
    """Process uploaded CSV and create Student/Application/Score records."""

    num_processed = 0

    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)

        for row in reader:
            if not row.get('Student Last Name'):
                continue

            # Create or update student
            student_id = row.get('Student ID') or f"{row['Student First Name']}{row['Student Last Name']}{num_processed}"

            student = Student.query.filter_by(student_id=student_id).first()
            if not student:
                student = Student(
                    student_id=student_id,
                    first_name=row.get('Student First Name', ''),
                    middle_name=row.get('Student Middle Name', ''),
                    last_name=row.get('Student Last Name', ''),
                    email=row.get('Email', ''),
                    gpa=Decimal(row.get('GPA', '0')) if row.get('GPA') else None,
                    program_level=row.get('Program Level', 'MS').upper(),
                    financial_need=row.get('Financial Need', '').upper()[0] if row.get('Financial Need') else None,
                    permanent_position_at=row.get('Permanent Position At', ''),
                    internship_2025_at=row.get('Internship 2025 At', ''),
                    internship_2026_at=row.get('Internship 2026 At', ''),
                )
                db.session.add(student)
                db.session.flush()

            # Create or update application
            application = Application.query.filter_by(student_id=student.id).first()
            if not application:
                application = Application(
                    student_id=student.id,
                    extracurricular_activities=row.get('Extracurricular Activities', ''),
                    college_activities=row.get('College Activities', ''),
                    work_experience=row.get('Work Experience', ''),
                )
                db.session.add(application)
            else:
                # Update existing application
                application.extracurricular_activities = row.get('Extracurricular Activities', '')
                application.college_activities = row.get('College Activities', '')
                application.work_experience = row.get('Work Experience', '')
            db.session.flush()

            # Score the application
            score_dict = score_student(student, application)

            # Create or update score
            score = Score.query.filter_by(student_id=student.id).first()
            if not score:
                score = Score(
                    student_id=student.id,
                    accounting_experience_score=Decimal(str(score_dict['accounting_experience_score'])),
                    work_experience_score=Decimal(str(score_dict['work_experience_score'])),
                    leadership_score=Decimal(str(score_dict['leadership_score'])),
                    total_score=Decimal(str(score_dict['total_score'])),
                    rubric_type=student.program_level,
                    scoring_notes=' | '.join(score_dict['notes'])
                )
                db.session.add(score)
            else:
                # Update existing score
                score.accounting_experience_score = Decimal(str(score_dict['accounting_experience_score']))
                score.work_experience_score = Decimal(str(score_dict['work_experience_score']))
                score.leadership_score = Decimal(str(score_dict['leadership_score']))
                score.total_score = Decimal(str(score_dict['total_score']))
                score.scoring_notes = ' | '.join(score_dict['notes'])

            # Create suggested award (or update if exists)
            eligible_scholarships = get_eligible_scholarships(student)
            if eligible_scholarships:
                scholarship, match_score = eligible_scholarships[0]

                # Check if award already exists
                existing_award = Award.query.filter_by(
                    student_id=student.id,
                    scholarship_id=scholarship.id
                ).first()

                if not existing_award:
                    award = Award(
                        student_id=student.id,
                        scholarship_id=scholarship.id,
                        amount=Decimal('5000'),  # Default amount
                        status='pending'
                    )
                    db.session.add(award)
                elif existing_award.status in ('rejected', 'notified'):
                    existing_award.status = 'pending'
                    existing_award.amount = Decimal('5000')
                    existing_award.approved_by = None
                    existing_award.approved_at = None

            num_processed += 1

    db.session.commit()
    return num_processed


def _process_applications_excel(filepath):
    """Process uploaded Excel file and create Student/Application/Score records."""
    
    # Read Excel file with pandas
    df = pd.read_excel(filepath)
    
    num_processed = 0
    
    for index, row in df.iterrows():
        # Skip if last name is empty
        if pd.isna(row.get('Student Last Name')) or str(row.get('Student Last Name', '')).strip() == '':
            continue
        
        # Create or update student
        student_id = str(row.get('Student ID', '')) if not pd.isna(row.get('Student ID')) else f"{row['Student First Name']}{row['Student Last Name']}{num_processed}"
        
        student = Student.query.filter_by(student_id=student_id).first()
        if not student:
            student = Student(
                student_id=student_id,
                first_name=str(row.get('Student First Name', '')).strip(),
                middle_name=str(row.get('Student Middle Name', '')).strip() if not pd.isna(row.get('Student Middle Name')) else '',
                last_name=str(row.get('Student Last Name', '')).strip(),
                email=str(row.get('Email', '')).strip() if not pd.isna(row.get('Email')) else '',
                gpa=Decimal(str(row.get('GPA', '0'))) if not pd.isna(row.get('GPA')) and str(row.get('GPA', '')).strip() else None,
                program_level=str(row.get('Program Level', 'MS')).upper().strip(),
                financial_need=str(row.get('Financial Need', '')).upper()[0] if not pd.isna(row.get('Financial Need')) and str(row.get('Financial Need', '')).strip() else None,
                permanent_position_at=str(row.get('Permanent Position At', '')).strip() if not pd.isna(row.get('Permanent Position At')) else '',
                internship_2025_at=str(row.get('Internship 2025 At', '')).strip() if not pd.isna(row.get('Internship 2025 At')) else '',
                internship_2026_at=str(row.get('Internship 2026 At', '')).strip() if not pd.isna(row.get('Internship 2026 At')) else '',
            )
            db.session.add(student)
            db.session.flush()
        
        # Create or update application
        application = Application.query.filter_by(student_id=student.id).first()
        if not application:
            application = Application(
                student_id=student.id,
                extracurricular_activities=str(row.get('Extracurricular Activities', '')).strip() if not pd.isna(row.get('Extracurricular Activities')) else '',
                college_activities=str(row.get('College Activities', '')).strip() if not pd.isna(row.get('College Activities')) else '',
                work_experience=str(row.get('Work Experience', '')).strip() if not pd.isna(row.get('Work Experience')) else '',
            )
            db.session.add(application)
        else:
            # Update existing application
            application.extracurricular_activities = str(row.get('Extracurricular Activities', '')).strip() if not pd.isna(row.get('Extracurricular Activities')) else ''
            application.college_activities = str(row.get('College Activities', '')).strip() if not pd.isna(row.get('College Activities')) else ''
            application.work_experience = str(row.get('Work Experience', '')).strip() if not pd.isna(row.get('Work Experience')) else ''
        
        db.session.flush()
        
        # Score the application
        score_dict = score_student(student, application)
        
        # Create or update score
        score = Score.query.filter_by(student_id=student.id).first()
        if not score:
            score = Score(
                student_id=student.id,
                accounting_experience_score=Decimal(str(score_dict['accounting_experience_score'])),
                work_experience_score=Decimal(str(score_dict['work_experience_score'])),
                leadership_score=Decimal(str(score_dict['leadership_score'])),
                total_score=Decimal(str(score_dict['total_score'])),
                rubric_type=student.program_level,
                scoring_notes=' | '.join(score_dict['notes'])
            )
            db.session.add(score)
        else:
            # Update existing score
            score.accounting_experience_score = Decimal(str(score_dict['accounting_experience_score']))
            score.work_experience_score = Decimal(str(score_dict['work_experience_score']))
            score.leadership_score = Decimal(str(score_dict['leadership_score']))
            score.total_score = Decimal(str(score_dict['total_score']))
            score.scoring_notes = ' | '.join(score_dict['notes'])
        
        # Create suggested award (or update if exists)
        eligible_scholarships = get_eligible_scholarships(student)
        if eligible_scholarships:
            scholarship, match_score = eligible_scholarships[0]
            
            # Check if award already exists
            existing_award = Award.query.filter_by(
                student_id=student.id,
                scholarship_id=scholarship.id
            ).first()
            
            if not existing_award:
                award = Award(
                    student_id=student.id,
                    scholarship_id=scholarship.id,
                    amount=Decimal('5000'),  # Default amount
                    status='pending'
                )
                db.session.add(award)
            elif existing_award.status in ('rejected', 'notified'):
                existing_award.status = 'pending'
                existing_award.amount = Decimal('5000')
                existing_award.approved_by = None
                existing_award.approved_at = None

        num_processed += 1

    db.session.commit()
    return num_processed


if __name__ == '__main__':
    with app.app_context():
        init_db()
        load_scholarships_from_excel()
        seed_sample_data()
    app.run(debug=app.config['DEBUG'], host='0.0.0.0', port=5000)
